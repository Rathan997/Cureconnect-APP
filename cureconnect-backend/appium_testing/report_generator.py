"""
Excel Report Generator — CureConnect Appium Android Tests
==========================================================
Generates a rich multi-sheet Excel workbook with:
  Sheet 1: Executive Summary (pass/fail counts, pass rate, pie-chart data)
  Sheet 2: All Test Results (per-test row with colour-coded status)
  Sheet 3: Failures & Errors (error messages + screenshot paths)
  Sheet 4: Test Coverage Matrix (feature × test-class breakdown)
  Sheet 5: Statistics & Analysis (per-class pass rate, duration)
  Sheet 6: Recommendations (standard QA action items)
  Sheet 7: Setup Guide (how to run the suite)
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
C_NAVY      = "1A237E"   # deep navy – header background
C_TEAL      = "00796B"   # teal – accent
C_GREEN_BG  = "C8E6C9"   # light green background
C_GREEN_FG  = "1B5E20"   # dark green text
C_RED_BG    = "FFCDD2"   # light red background
C_RED_FG    = "B71C1C"   # dark red text
C_YELLOW_BG = "FFF9C4"   # light yellow background
C_YELLOW_FG = "F57F17"   # dark yellow text
C_BLUE_BG   = "BBDEFB"   # light blue background
C_BLUE_FG   = "0D47A1"   # dark blue text
C_GREY_BG   = "F5F5F5"   # alternating row grey
C_WHITE     = "FFFFFF"
C_BLACK     = "000000"


def _fill(hex_colour):
    return PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")


def _font(bold=False, size=11, colour=C_BLACK, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=colour, italic=italic)


def _border():
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Main class
# ---------------------------------------------------------------------------

class AppiumExcelReportGenerator:
    """Generates a comprehensive Excel test report for Appium Android tests."""

    def __init__(self, report_dir=None):
        if report_dir is None:
            import config
            report_dir = config.REPORTS_DIR
        self.report_dir = Path(report_dir)
        self.report_dir.mkdir(parents=True, exist_ok=True)

        self.wb = Workbook()
        self.wb.remove(self.wb.active)   # remove default blank sheet

        self.results   = []   # list[dict] – one per test
        self.issues    = []   # failed / errored tests
        self.stats = {"total": 0, "passed": 0, "failed": 0, "skipped": 0,
                      "total_duration": 0.0}

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def add_result(self, test_id, test_name, test_class, status,
                   duration=0.0, error_message=None, screenshot=None,
                   feature=None):
        """Register one test result."""
        rec = {
            "test_id":      test_id,
            "test_name":    test_name,
            "test_class":   test_class,
            "feature":      feature or test_class,
            "status":       status.upper(),
            "duration":     round(float(duration), 2),
            "error_message": error_message,
            "screenshot":   screenshot,
            "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        self.results.append(rec)
        self.stats["total"] += 1
        self.stats["total_duration"] += rec["duration"]
        if rec["status"] == "PASSED":
            self.stats["passed"] += 1
        elif rec["status"] == "FAILED":
            self.stats["failed"] += 1
            self.issues.append(rec)
        elif rec["status"] == "SKIPPED":
            self.stats["skipped"] += 1
        else:
            self.stats["failed"] += 1
            self.issues.append(rec)
        logger.debug("Result added: %s – %s", test_id, status)

    def generate(self):
        """Build and save the Excel workbook. Returns the Path of the saved file."""
        logger.info("Generating Appium Excel report …")
        self._sheet_summary()
        self._sheet_results()
        self._sheet_failures()
        self._sheet_coverage()
        self._sheet_statistics()
        self._sheet_recommendations()
        self._sheet_setup_guide()

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = self.report_dir / f"Appium_Test_Report_{ts}.xlsx"
        self.wb.save(str(out))
        logger.info("Report saved → %s", out)
        return out

    # ------------------------------------------------------------------
    # Sheet helpers
    # ------------------------------------------------------------------

    def _write_header_row(self, ws, row, headers, bg=C_NAVY):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font      = _font(bold=True, colour=C_WHITE, size=11)
            c.fill      = _fill(bg)
            c.alignment = _center()
            c.border    = _border()

    def _pass_rate(self):
        return round(self.stats["passed"] / self.stats["total"] * 100, 1) if self.stats["total"] else 0

    def _status_style(self, cell, status):
        if status == "PASSED":
            cell.fill = _fill(C_GREEN_BG)
            cell.font = _font(bold=True, colour=C_GREEN_FG)
        elif status == "FAILED":
            cell.fill = _fill(C_RED_BG)
            cell.font = _font(bold=True, colour=C_RED_FG)
        elif status == "SKIPPED":
            cell.fill = _fill(C_YELLOW_BG)
            cell.font = _font(bold=True, colour=C_YELLOW_FG)
        cell.alignment = _center()
        cell.border    = _border()

    # ------------------------------------------------------------------
    # Sheet 1 – Summary
    # ------------------------------------------------------------------

    def _sheet_summary(self):
        ws = self.wb.create_sheet("📊 Summary", 0)
        ws.sheet_view.showGridLines = False

        # ── Title banner ──
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value     = "🤖  CURECONNECT  ·  APPIUM ANDROID TEST REPORT"
        t.font      = _font(bold=True, size=18, colour=C_WHITE)
        t.fill      = _fill(C_NAVY)
        t.alignment = _center()

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value     = f"Generated: {datetime.now().strftime('%A, %d %B %Y  %H:%M:%S')}"
        s.font      = _font(italic=True, size=10, colour=C_TEAL)
        s.alignment = _center()
        ws.row_dimensions[1].height = 38
        ws.row_dimensions[2].height = 22

        # ── KPI cards (row 4-5 for labels, 6-7 for values) ──
        kpis = [
            ("Total Tests",  self.stats["total"],     C_BLUE_BG,   C_BLUE_FG,   "A"),
            ("Passed ✅",    self.stats["passed"],    C_GREEN_BG,  C_GREEN_FG,  "B"),
            ("Failed ❌",    self.stats["failed"],    C_RED_BG,    C_RED_FG,    "C"),
            ("Skipped ⏭",   self.stats["skipped"],   C_YELLOW_BG, C_YELLOW_FG, "D"),
            ("Pass Rate",    f"{self._pass_rate()}%", C_BLUE_BG,   C_BLUE_FG,   "E"),
            ("Duration (s)", round(self.stats["total_duration"], 1), C_GREY_BG, "000000", "F"),
        ]

        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 40

        for label, value, bg, fg, col in kpis:
            lbl_cell = ws[f"{col}4"]
            val_cell = ws[f"{col}5"]

            lbl_cell.value     = label
            lbl_cell.font      = _font(bold=True, size=10, colour="555555")
            lbl_cell.alignment = _center()
            lbl_cell.fill      = _fill(bg)
            lbl_cell.border    = _border()

            val_cell.value     = value
            val_cell.font      = _font(bold=True, size=24, colour=fg)
            val_cell.alignment = _center()
            val_cell.fill      = _fill(bg)
            val_cell.border    = _border()

        # ── Result breakdown table ──
        row = 8
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"].value     = "TEST RESULT BREAKDOWN"
        ws[f"A{row}"].font      = _font(bold=True, size=12, colour=C_WHITE)
        ws[f"A{row}"].fill      = _fill(C_TEAL)
        ws[f"A{row}"].alignment = _center()

        row += 1
        self._write_header_row(ws, row, ["#", "Test Name", "Feature / Class", "Status", "Duration (s)", "Timestamp"], C_TEAL)

        for idx, r in enumerate(self.results, 1):
            row += 1
            bg = C_GREY_BG if idx % 2 == 0 else C_WHITE
            cells_data = [
                (idx,            _center()),
                (r["test_name"], _left()),
                (r["feature"],   _left()),
                (r["status"],    _center()),
                (r["duration"],  _center()),
                (r["timestamp"], _center()),
            ]
            for col, (val, align) in enumerate(cells_data, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.alignment = align
                c.border    = _border()
                if col != 4:
                    c.fill = _fill(bg)
            # colour the status cell
            self._status_style(ws.cell(row=row, column=4), r["status"])

        # ── Column widths ──
        for col_w in [("A", 6), ("B", 40), ("C", 28), ("D", 14), ("E", 14), ("F", 22)]:
            ws.column_dimensions[col_w[0]].width = col_w[1]

    # ------------------------------------------------------------------
    # Sheet 2 – Full Results
    # ------------------------------------------------------------------

    def _sheet_results(self):
        ws = self.wb.create_sheet("✅ Test Results")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:G1")
        ws["A1"].value     = "ALL TEST RESULTS — CURECONNECT APPIUM SUITE"
        ws["A1"].font      = _font(bold=True, size=14, colour=C_WHITE)
        ws["A1"].fill      = _fill(C_NAVY)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 32

        headers = ["Test ID", "Test Name", "Test Class", "Feature", "Status", "Duration (s)", "Timestamp"]
        self._write_header_row(ws, 2, headers)

        for idx, r in enumerate(self.results, 3):
            bg = C_GREY_BG if idx % 2 == 0 else C_WHITE
            row_data = [r["test_id"], r["test_name"], r["test_class"],
                        r["feature"], r["status"], r["duration"], r["timestamp"]]
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=idx, column=col, value=val)
                c.border    = _border()
                c.alignment = _center() if col in (1, 5, 6, 7) else _left()
                c.fill      = _fill(bg)
            self._status_style(ws.cell(row=idx, column=5), r["status"])

        for col, w in zip("ABCDEFG", [12, 40, 28, 24, 12, 14, 22]):
            ws.column_dimensions[col].width = w

    # ------------------------------------------------------------------
    # Sheet 3 – Failures
    # ------------------------------------------------------------------

    def _sheet_failures(self):
        ws = self.wb.create_sheet("❌ Failures")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:E1")
        ws["A1"].value     = "FAILED TESTS — DETAILS & SCREENSHOTS"
        ws["A1"].font      = _font(bold=True, size=14, colour=C_WHITE)
        ws["A1"].fill      = _fill("B71C1C")
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 32

        if not self.issues:
            ws["A3"].value = "🎉  All tests passed!  No failures recorded."
            ws["A3"].font  = _font(bold=True, size=13, colour=C_GREEN_FG)
            return

        headers = ["Test ID", "Test Name", "Test Class", "Error Message", "Screenshot Path"]
        self._write_header_row(ws, 2, headers, "B71C1C")

        for idx, r in enumerate(self.issues, 3):
            bg = C_RED_BG
            data = [r["test_id"], r["test_name"], r["test_class"],
                    r.get("error_message") or "—", r.get("screenshot") or "—"]
            for col, val in enumerate(data, 1):
                c = ws.cell(row=idx, column=col, value=val)
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = _left()
                c.font      = _font(size=10)

        for col, w in zip("ABCDE", [12, 36, 26, 60, 40]):
            ws.column_dimensions[col].width = w
        ws.row_dimensions[2].height = 22

    # ------------------------------------------------------------------
    # Sheet 4 – Coverage Matrix
    # ------------------------------------------------------------------

    def _sheet_coverage(self):
        ws = self.wb.create_sheet("🗺 Coverage")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:D1")
        ws["A1"].value     = "FEATURE COVERAGE MATRIX"
        ws["A1"].font      = _font(bold=True, size=14, colour=C_WHITE)
        ws["A1"].fill      = _fill(C_NAVY)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 32

        # Build per-class stats
        classes = {}
        for r in self.results:
            cls = r["test_class"]
            if cls not in classes:
                classes[cls] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0}
            classes[cls]["total"] += 1
            if r["status"] == "PASSED":
                classes[cls]["passed"] += 1
            elif r["status"] == "FAILED":
                classes[cls]["failed"] += 1
            else:
                classes[cls]["skipped"] += 1

        headers = ["Test Class / Feature Area", "Total", "Passed", "Failed", "Pass Rate"]
        self._write_header_row(ws, 3, headers, C_TEAL)

        for row_idx, (cls, s) in enumerate(classes.items(), 4):
            rate = f"{round(s['passed'] / s['total'] * 100, 1)}%" if s["total"] else "N/A"
            bg = C_GREEN_BG if s["failed"] == 0 else C_RED_BG
            for col, val in enumerate([cls, s["total"], s["passed"], s["failed"], rate], 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = _left() if col == 1 else _center()
                c.font      = _font(size=10)

        for col, w in zip("ABCDE", [34, 10, 10, 10, 12]):
            ws.column_dimensions[col].width = w

        # ── App features tested table ──
        row = len(classes) + 6
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"].value     = "APP FEATURE COVERAGE CHECKLIST"
        ws[f"A{row}"].font      = _font(bold=True, size=12, colour=C_WHITE)
        ws[f"A{row}"].fill      = _fill(C_TEAL)
        ws[f"A{row}"].alignment = _center()

        features = [
            ("Authentication",       "Login / Logout / Forgot Password",         "✅ Covered"),
            ("Onboarding",           "Splash → Onboarding → Login",              "✅ Covered"),
            ("Navigation",           "Bottom Tab Bar + Quick Actions",           "✅ Covered"),
            ("Doctors Screen",       "List, Search, Filter, Detail view",        "✅ Covered"),
            ("Appointments",         "View, Tab filter, Cancel appointment",     "✅ Covered"),
            ("Symptom Checker",      "AI symptom selection and analysis",        "✅ Covered"),
            ("Medicine Scanner",     "Add medicine manually + list view",        "✅ Covered"),
            ("Family Management",    "Add/View/Check-in family members",         "✅ Covered"),
            ("Health Dashboard",     "Score ring, stats, weekly chart",          "✅ Covered"),
            ("Emergency SOS",        "Emergency contacts + call buttons",        "✅ Covered"),
            ("AI Chat",              "Chat interface and message flow",          "✅ Covered"),
            ("Profile",              "View profile info and edit profile",       "✅ Covered"),
            ("End-to-End Journey",   "Login → Features → Logout full flow",     "✅ Covered"),
            ("Negative / Edge Cases","Invalid login, empty form, back nav",      "✅ Covered"),
        ]

        row += 1
        self._write_header_row(ws, row, ["Feature", "Description", "Status"], C_NAVY)
        for f_row, (feat, desc, status) in enumerate(features, row + 1):
            bg = C_GREEN_BG if "✅" in status else C_YELLOW_BG
            for col, val in enumerate([feat, desc, status], 1):
                c = ws.cell(row=f_row, column=col, value=val)
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = _left() if col < 3 else _center()
                c.font      = _font(size=10)

    # ------------------------------------------------------------------
    # Sheet 5 – Statistics
    # ------------------------------------------------------------------

    def _sheet_statistics(self):
        ws = self.wb.create_sheet("📈 Statistics")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:C1")
        ws["A1"].value     = "STATISTICS & ANALYSIS"
        ws["A1"].font      = _font(bold=True, size=14, colour=C_WHITE)
        ws["A1"].fill      = _fill(C_NAVY)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 32

        stats_rows = [
            ("Metric",              "Value",          "Notes"),
            ("Total Tests",         self.stats["total"],    "All test methods executed"),
            ("Passed",              self.stats["passed"],   "Tests that completed successfully"),
            ("Failed",              self.stats["failed"],   "Tests that raised AssertionError / Exception"),
            ("Skipped",             self.stats["skipped"],  "Tests skipped via @unittest.skip"),
            ("Pass Rate (%)",       f"{self._pass_rate()}%","passed / total × 100"),
            ("Fail Rate (%)",       f"{round(100 - self._pass_rate(), 1)}%", "failed / total × 100"),
            ("Total Duration (s)",  round(self.stats["total_duration"], 2), "Sum of all test durations"),
            ("Avg Duration (s)",
             round(self.stats["total_duration"] / max(self.stats["total"], 1), 2),
             "Total duration / total tests"),
        ]

        self._write_header_row(ws, 3, ["Metric", "Value", "Notes"], C_TEAL)
        for row_idx, (m, v, n) in enumerate(stats_rows[1:], 4):
            bg = C_GREY_BG if row_idx % 2 == 0 else C_WHITE
            for col, val in enumerate([m, v, n], 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = _left() if col != 2 else _center()
                c.font      = _font(size=11, bold=(col == 1))

        for col, w in zip("ABC", [26, 14, 46]):
            ws.column_dimensions[col].width = w

        # ── Bar chart (passed vs failed per class) ──
        classes = {}
        for r in self.results:
            cls = r["test_class"]
            classes.setdefault(cls, {"passed": 0, "failed": 0})
            if r["status"] == "PASSED":
                classes[cls]["passed"] += 1
            elif r["status"] == "FAILED":
                classes[cls]["failed"] += 1

        chart_start_row = 15
        ws.cell(row=chart_start_row, column=1, value="Class").font = _font(bold=True)
        ws.cell(row=chart_start_row, column=2, value="Passed").font = _font(bold=True)
        ws.cell(row=chart_start_row, column=3, value="Failed").font = _font(bold=True)

        for i, (cls, v) in enumerate(classes.items(), chart_start_row + 1):
            ws.cell(row=i, column=1, value=cls)
            ws.cell(row=i, column=2, value=v["passed"])
            ws.cell(row=i, column=3, value=v["failed"])

        if classes:
            data_rows = len(classes)
            chart = BarChart()
            chart.type = "col"
            chart.title = "Passed vs Failed per Test Class"
            chart.y_axis.title = "Tests"
            chart.x_axis.title = "Test Class"
            chart.width  = 22
            chart.height = 14

            data_ref = Reference(ws, min_col=2, max_col=3,
                                 min_row=chart_start_row,
                                 max_row=chart_start_row + data_rows)
            cats_ref = Reference(ws, min_col=1,
                                 min_row=chart_start_row + 1,
                                 max_row=chart_start_row + data_rows)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "E4")

    # ------------------------------------------------------------------
    # Sheet 6 – Recommendations
    # ------------------------------------------------------------------

    def _sheet_recommendations(self):
        ws = self.wb.create_sheet("💡 Recommendations")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:C1")
        ws["A1"].value     = "QA RECOMMENDATIONS & ACTION ITEMS"
        ws["A1"].font      = _font(bold=True, size=14, colour=C_WHITE)
        ws["A1"].fill      = _fill(C_NAVY)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 32

        recs = [
            ("Priority", "Category", "Recommendation"),
            ("🔴 High",   "Failed Tests",
             "Investigate and fix all FAILED test cases before the next release."),
            ("🔴 High",   "Locators",
             "Re-run tests after any UI change; update accessibility testID values accordingly."),
            ("🟡 Medium", "Coverage",
             "Add negative tests: network timeout, empty states, permission denial flows."),
            ("🟡 Medium", "Stability",
             "Add retry logic for flaky network-dependent steps (symptom checker API, login)."),
            ("🟡 Medium", "Test Data",
             "Externalize test credentials and medicine data to a separate .env or JSON file."),
            ("🟢 Low",    "CI/CD",
             "Integrate this suite into GitHub Actions / Bitrise for automated regression."),
            ("🟢 Low",    "Performance",
             "Capture and assert response times for API-heavy screens (Doctors, Symptom)."),
            ("🟢 Low",    "Parallelism",
             "Use pytest-xdist or Appium grid to run test classes in parallel."),
            ("🟢 Low",    "Reporting",
             "Schedule daily test runs and email this Excel report to the team automatically."),
            ("🟢 Low",    "Accessibility",
             "Ensure all interactive elements have unique testID values for reliable automation."),
        ]

        self._write_header_row(ws, 3, recs[0], C_TEAL)
        for row_idx, (pri, cat, rec) in enumerate(recs[1:], 4):
            bg = (C_RED_BG if "🔴" in pri else
                  C_YELLOW_BG if "🟡" in pri else C_GREEN_BG)
            for col, val in enumerate([pri, cat, rec], 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = _center() if col < 3 else _left()
                c.font      = _font(size=10)

        for col, w in zip("ABC", [14, 20, 70]):
            ws.column_dimensions[col].width = w

    # ------------------------------------------------------------------
    # Sheet 7 – Setup Guide
    # ------------------------------------------------------------------

    def _sheet_setup_guide(self):
        ws = self.wb.create_sheet("🚀 Setup Guide")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:B1")
        ws["A1"].value     = "HOW TO RUN THE APPIUM TEST SUITE"
        ws["A1"].font      = _font(bold=True, size=14, colour=C_WHITE)
        ws["A1"].fill      = _fill(C_NAVY)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 32

        steps = [
            ("Step", "Command / Action"),
            ("1. Install Appium 2",        "npm install -g appium@next"),
            ("2. Install UiAutomator2",    "appium driver install uiautomator2"),
            ("3. Install Python deps",     "pip install -r requirements.txt"),
            ("4. Build APK",               "eas build -p android --profile preview   OR   expo prebuild"),
            ("5. Connect Android device",  "adb devices   (ensure device shows as 'device')"),
            ("6. Start Appium server",     "appium --port 4723"),
            ("7. Run the full suite",      "python run_tests.py"),
            ("8. Run specific class",      "python run_tests.py --class Auth"),
            ("9. Open Excel report",       "Reports are saved to:  appium_testing/reports/"),
            ("10. View screenshots",       "Screenshots are in:    appium_testing/reports/screenshots/"),
        ]

        self._write_header_row(ws, 3, ["Step", "Command / Action"], C_TEAL)
        for row_idx, (step, cmd) in enumerate(steps[1:], 4):
            bg = C_GREY_BG if row_idx % 2 == 0 else C_WHITE
            for col, val in enumerate([step, cmd], 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = _left()
                c.font      = _font(size=10, bold=(col == 1))

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 70

        # Environment variables note
        note_row = len(steps) + 5
        ws.merge_cells(f"A{note_row}:B{note_row}")
        ws[f"A{note_row}"].value = (
            "ENV VARS:  APPIUM_HOST, APPIUM_PORT, DEVICE_NAME, DEVICE_SERIAL, "
            "PLATFORM_VERSION, APK_PATH — override via OS environment variables."
        )
        ws[f"A{note_row}"].font      = _font(italic=True, size=9, colour="555555")
        ws[f"A{note_row}"].alignment = _left()
