"""
Excel Report Generator for Selenium Tests
Generates comprehensive test reports with formatting
"""

import logging
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate Excel test reports"""
    
    # Colors
    COLOR_GREEN = "C6EFCE"  # Green
    COLOR_RED = "FFC7CE"    # Red
    COLOR_YELLOW = "FFEB9C" # Yellow
    COLOR_BLUE = "BDD7EE"   # Blue
    COLOR_HEADER = "366092" # Dark Blue
    
    def __init__(self, report_dir=None):
        """Initialize report generator"""
        if report_dir is None:
            from config import REPORTS_DIR
            report_dir = REPORTS_DIR
        
        self.report_dir = Path(report_dir)
        self.report_dir.mkdir(exist_ok=True)
        
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        self.test_results = []
        self.test_issues = []
        self.statistics = {
            "total": 0,
            "passed": 0,
            "failed": 0,
            "skipped": 0
        }
        
        logger.info("Excel Report Generator initialized")
    
    def add_test_result(self, test_id, test_name, test_class, status, 
                        duration=0, error_message=None, screenshot=None):
        """Add test result"""
        result = {
            "test_id": test_id,
            "test_name": test_name,
            "test_class": test_class,
            "status": status,
            "duration": duration,
            "error_message": error_message,
            "screenshot": screenshot,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        self.test_results.append(result)
        
        # Update statistics
        self.statistics["total"] += 1
        if status == "PASSED":
            self.statistics["passed"] += 1
        elif status == "FAILED":
            self.statistics["failed"] += 1
            self.test_issues.append(result)
        elif status == "SKIPPED":
            self.statistics["skipped"] += 1
        
        logger.debug(f"Test result added: {test_id} - {status}")
    
    def generate_report(self):
        """Generate Excel report with all sheets"""
        try:
            logger.info("Generating Excel report...")
            
            self._add_summary_sheet()
            self._add_test_results_sheet()
            self._add_issues_sheet()
            self._add_recommendations_sheet()
            self._add_statistics_sheet()
            
            # Save report
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_file = self.report_dir / f"Selenium_Test_Report_{timestamp}.xlsx"
            self.workbook.save(str(report_file))
            
            logger.info(f"✓ Report generated: {report_file}")
            return report_file
            
        except Exception as e:
            logger.error(f"✗ Failed to generate report: {e}")
            raise
    
    def _add_summary_sheet(self):
        """Add summary sheet"""
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Title
        title_cell = ws["A1"]
        title_cell.value = "SELENIUM TEST EXECUTION REPORT"
        title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        ws.merge_cells("A1:D1")
        
        # Timestamp
        ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(italic=True)
        
        # Empty row
        ws["A3"].value = ""
        
        # Statistics boxes
        row = 4
        
        # Total Tests
        ws[f"A{row}"].value = "Total Tests"
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLOR_BLUE, end_color=self.COLOR_BLUE, fill_type="solid")
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
        ws[f"B{row}"].value = self.statistics["total"]
        ws[f"B{row}"].font = Font(size=12, bold=True)
        row += 1
        
        # Passed
        ws[f"A{row}"].value = "Passed"
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLOR_GREEN, end_color=self.COLOR_GREEN, fill_type="solid")
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].value = self.statistics["passed"]
        ws[f"B{row}"].font = Font(size=12, bold=True, color="008000")
        row += 1
        
        # Failed
        ws[f"A{row}"].value = "Failed"
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLOR_RED, end_color=self.COLOR_RED, fill_type="solid")
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].value = self.statistics["failed"]
        ws[f"B{row}"].font = Font(size=12, bold=True, color="FF0000")
        row += 1
        
        # Skipped
        ws[f"A{row}"].value = "Skipped"
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLOR_YELLOW, end_color=self.COLOR_YELLOW, fill_type="solid")
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].value = self.statistics["skipped"]
        ws[f"B{row}"].font = Font(size=12, bold=True, color="FF9900")
        row += 1
        
        # Pass Rate
        row += 1
        ws[f"A{row}"].value = "Pass Rate"
        ws[f"A{row}"].font = Font(bold=True, size=11)
        if self.statistics["total"] > 0:
            pass_rate = (self.statistics["passed"] / self.statistics["total"]) * 100
            ws[f"B{row}"].value = f"{pass_rate:.1f}%"
        else:
            ws[f"B{row}"].value = "N/A"
        ws[f"B{row}"].font = Font(bold=True, size=11)
        
        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        
        logger.debug("Summary sheet added")
    
    def _add_test_results_sheet(self):
        """Add test results sheet"""
        ws = self.workbook.create_sheet("Test Results")
        
        # Headers
        headers = ["Test ID", "Test Name", "Test Class", "Status", "Duration (s)", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add test results
        for row_idx, result in enumerate(self.test_results, 2):
            ws.cell(row=row_idx, column=1).value = result["test_id"]
            ws.cell(row=row_idx, column=2).value = result["test_name"]
            ws.cell(row=row_idx, column=3).value = result["test_class"]
            
            # Status with color coding
            status_cell = ws.cell(row=row_idx, column=4)
            status_cell.value = result["status"]
            if result["status"] == "PASSED":
                status_cell.fill = PatternFill(start_color=self.COLOR_GREEN, end_color=self.COLOR_GREEN, fill_type="solid")
                status_cell.font = Font(color="008000", bold=True)
            elif result["status"] == "FAILED":
                status_cell.fill = PatternFill(start_color=self.COLOR_RED, end_color=self.COLOR_RED, fill_type="solid")
                status_cell.font = Font(color="FF0000", bold=True)
            else:
                status_cell.fill = PatternFill(start_color=self.COLOR_YELLOW, end_color=self.COLOR_YELLOW, fill_type="solid")
            
            ws.cell(row=row_idx, column=5).value = result.get("duration", 0)
            ws.cell(row=row_idx, column=6).value = result.get("timestamp", "")
            
            # Center alignment for status and duration
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")
        
        # Set column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 20
        
        logger.debug(f"Test results sheet added with {len(self.test_results)} results")
    
    def _add_issues_sheet(self):
        """Add issues/failures sheet"""
        ws = self.workbook.create_sheet("Issues")
        
        if not self.test_issues:
            ws["A1"].value = "No issues found - All tests passed!"
            ws["A1"].font = Font(bold=True, size=12, color="008000")
            return
        
        # Headers
        headers = ["Test ID", "Test Name", "Status", "Error Message", "Screenshot"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        
        # Add issues
        for row_idx, issue in enumerate(self.test_issues, 2):
            ws.cell(row=row_idx, column=1).value = issue["test_id"]
            ws.cell(row=row_idx, column=2).value = issue["test_name"]
            
            status_cell = ws.cell(row=row_idx, column=3)
            status_cell.value = issue["status"]
            status_cell.fill = PatternFill(start_color=self.COLOR_RED, end_color=self.COLOR_RED, fill_type="solid")
            status_cell.font = Font(bold=True, color="FFFFFF")
            
            error_cell = ws.cell(row=row_idx, column=4)
            error_cell.value = issue.get("error_message", "Unknown error")
            error_cell.alignment = Alignment(wrap_text=True)
            
            ws.cell(row=row_idx, column=5).value = issue.get("screenshot", "N/A")
        
        # Set column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 30
        
        logger.debug(f"Issues sheet added with {len(self.test_issues)} issues")
    
    def _add_recommendations_sheet(self):
        """Add recommendations sheet"""
        ws = self.workbook.create_sheet("Recommendations")
        
        # Title
        ws["A1"].value = "TEST EXECUTION RECOMMENDATIONS"
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        ws.merge_cells("A1:B1")
        
        # Recommendations
        recommendations = [
            ("Test Coverage", "Ensure all critical features are covered by tests"),
            ("Performance", "Review and optimize slow-running tests"),
            ("Error Handling", "Investigate and fix failed tests"),
            ("Maintenance", "Update locators and test data regularly"),
            ("CI/CD Integration", "Integrate tests into continuous integration pipeline"),
            ("Documentation", "Keep test documentation up-to-date"),
            ("Best Practices", "Follow Page Object Model pattern for new tests"),
            ("Scalability", "Consider parallel test execution for faster results"),
        ]
        
        row = 3
        for category, recommendation in recommendations:
            ws[f"A{row}"].value = category
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].value = recommendation
            ws[f"B{row}"].alignment = Alignment(wrap_text=True)
            row += 1
        
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60
        
        logger.debug("Recommendations sheet added")
    
    def _add_statistics_sheet(self):
        """Add statistics and analysis sheet"""
        ws = self.workbook.create_sheet("Statistics")
        
        # Title
        ws["A1"].value = "TEST STATISTICS & ANALYSIS"
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        ws.merge_cells("A1:B1")
        
        # Test execution info
        row = 3
        
        stats_data = [
            ("Total Tests", self.statistics["total"]),
            ("Passed", self.statistics["passed"]),
            ("Failed", self.statistics["failed"]),
            ("Skipped", self.statistics["skipped"]),
            ("Pass Rate (%)", self._calculate_pass_rate()),
            ("Fail Rate (%)", self._calculate_fail_rate()),
        ]
        
        for label, value in stats_data:
            ws[f"A{row}"].value = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].value = value
            row += 1
        
        # Test class breakdown
        row += 2
        ws[f"A{row}"].value = "TESTS BY CLASS"
        ws[f"A{row}"].font = Font(bold=True, underline="single")
        row += 1
        
        test_classes = {}
        for result in self.test_results:
            test_class = result["test_class"]
            if test_class not in test_classes:
                test_classes[test_class] = {"passed": 0, "failed": 0, "total": 0}
            test_classes[test_class]["total"] += 1
            if result["status"] == "PASSED":
                test_classes[test_class]["passed"] += 1
            elif result["status"] == "FAILED":
                test_classes[test_class]["failed"] += 1
        
        for test_class, stats in test_classes.items():
            ws[f"A{row}"].value = test_class
            ws[f"B{row}"].value = f"Passed: {stats['passed']}/{stats['total']}"
            row += 1
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        
        logger.debug("Statistics sheet added")
    
    def _calculate_pass_rate(self):
        """Calculate pass rate percentage"""
        if self.statistics["total"] == 0:
            return 0
        return round((self.statistics["passed"] / self.statistics["total"]) * 100, 2)
    
    def _calculate_fail_rate(self):
        """Calculate fail rate percentage"""
        if self.statistics["total"] == 0:
            return 0
        return round((self.statistics["failed"] / self.statistics["total"]) * 100, 2)


def generate_sample_report():
    """Generate sample report for testing"""
    from config import REPORTS_DIR
    
    generator = ExcelReportGenerator(REPORTS_DIR)
    
    # Add sample results
    generator.add_test_result("T001", "Login Test", "TestAuthentication", "PASSED", 2.5)
    generator.add_test_result("T002", "Dashboard Test", "TestDashboard", "PASSED", 3.2)
    generator.add_test_result("T003", "Doctors Search", "TestDoctors", "PASSED", 4.1)
    generator.add_test_result("T004", "Medicine Test", "TestMedicines", "FAILED", 5.0, "Element not found")
    
    return generator.generate_report()


if __name__ == "__main__":
    print("Generating sample report...")
    report_path = generate_sample_report()
    print(f"Report saved to: {report_path}")
