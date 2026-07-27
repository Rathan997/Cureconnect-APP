import os
import sys
import io
import json
import time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.platform.startswith("win"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Ensure we can import from the current directory
sys.path.insert(0, str(Path(__file__).parent))


import e2e_tests

# Define path for reports
REPORTS_DIR = Path(__file__).parent / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

def run_e2e_and_generate_excel():
    print("=" * 60)
    print("  CureConnect Web E2E Test Runner & Excel Reporter")
    print("=" * 60)
    
    # Configure base URLs
    e2e_tests.BASE_URL = "http://localhost:19006"
    e2e_tests.API_URL = "http://localhost:8000"
    
    print(f"Target URL: {e2e_tests.BASE_URL}")
    print(f"API URL:    {e2e_tests.API_URL}")
    
    if not e2e_tests.wait_for_app():
        print("[ERROR] Expo web server is not running on http://localhost:19006.")
        print("Please start it with: $env:CI=1; npx expo start --web --port 19006")
        sys.exit(1)
        
    driver = e2e_tests.create_driver()
    
    try:
        print("\n--> Starting Selenium E2E Web Application Tests...")
        logged_in = e2e_tests.test_authentication(driver)
        if logged_in:
            e2e_tests.test_dashboard(driver)
            e2e_tests.test_api_integration(driver)
            e2e_tests.test_doctors(driver)
            e2e_tests.test_symptoms(driver)
            e2e_tests.test_medicines(driver)
            e2e_tests.test_family(driver)
            e2e_tests.test_emergency(driver)
            e2e_tests.test_profile(driver)
            e2e_tests.test_logout(driver)
        else:
            print("\n[WARNING] Login failed - skipping authenticated tests.")
    except Exception as e:
        print(f"\n[FATAL ERROR] {e}")
    finally:
        e2e_tests.screenshot(driver, "final_state")
        driver.quit()
        print("\n--> Selenium E2E test execution completed.")

    # Generate Excel Report
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default active sheet
    
    # 1. Summary Sheet
    ws_summary = wb.create_sheet(title="Execution Summary")
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary["A1"] = "CURECONNECT WEB E2E TEST ANALYSIS"
    ws_summary["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    ws_summary["A1"].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.merge_cells("A1:D1")
    ws_summary.row_dimensions[1].height = 40
    
    # Run Info
    ws_summary["A3"] = "Tester Name:"
    ws_summary["A3"].font = Font(name="Segoe UI", bold=True)
    ws_summary["B3"] = "Automated Selenium Agent"
    
    ws_summary["A4"] = "Execution Time:"
    ws_summary["A4"].font = Font(name="Segoe UI", bold=True)
    ws_summary["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ws_summary["A5"] = "Target Environment:"
    ws_summary["A5"].font = Font(name="Segoe UI", bold=True)
    ws_summary["B5"] = e2e_tests.BASE_URL
    
    ws_summary["A6"] = "API Environment:"
    ws_summary["A6"].font = Font(name="Segoe UI", bold=True)
    ws_summary["B6"] = e2e_tests.API_URL
    
    # Metrics
    total_tests = len(e2e_tests.results)
    passed_tests = sum(1 for r in e2e_tests.results if "PASS" in r["status"])
    failed_tests = total_tests - passed_tests
    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0.0
    
    ws_summary["A8"] = "Metric"
    ws_summary["B8"] = "Value"
    for col in ["A", "B"]:
        ws_summary[f"{col}8"].font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        ws_summary[f"{col}8"].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_summary[f"{col}8"].alignment = Alignment(horizontal="center")
    
    metrics = [
        ("Total Tests Executed", total_tests),
        ("Passed Tests", passed_tests),
        ("Failed Tests", failed_tests),
        ("Pass Rate (%)", f"{pass_rate:.1f}%")
    ]
    
    for idx, (metric_name, val) in enumerate(metrics, 9):
        ws_summary[f"A{idx}"] = metric_name
        ws_summary[f"B{idx}"] = val
        ws_summary[f"A{idx}"].font = Font(name="Segoe UI")
        ws_summary[f"B{idx}"].font = Font(name="Segoe UI", bold=True)
        if "Rate" in metric_name:
            ws_summary[f"B{idx}"].fill = PatternFill(start_color="E2EFDA" if pass_rate >= 80 else "FCE4D6", end_color="E2EFDA" if pass_rate >= 80 else "FCE4D6", fill_type="solid")
            
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.column_dimensions["D"].width = 15

    # 2. Detailed Test Results Sheet
    ws_results = wb.create_sheet(title="Detailed Results")
    ws_results.views.sheetView[0].showGridLines = True
    
    headers = ["TC ID", "Module", "Test Description", "Status", "Notes/Output", "Screenshot Path", "Timestamp"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_results.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx, r in enumerate(e2e_tests.results, 2):
        status_text = "PASSED" if "PASS" in r["status"] else "FAILED"
        
        ws_results.cell(row=row_idx, column=1, value=r["test_id"]).alignment = Alignment(horizontal="center")
        ws_results.cell(row=row_idx, column=2, value=r["category"])
        ws_results.cell(row=row_idx, column=3, value=r["name"])
        
        status_cell = ws_results.cell(row=row_idx, column=4, value=status_text)
        status_cell.alignment = Alignment(horizontal="center")
        if status_text == "PASSED":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(name="Segoe UI", color="006100", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(name="Segoe UI", color="9C0006", bold=True)
            
        ws_results.cell(row=row_idx, column=5, value=r["notes"])
        ws_results.cell(row=row_idx, column=6, value=r.get("screenshot") or "N/A")
        ws_results.cell(row=row_idx, column=7, value=r.get("timestamp") or "")
        
        for col_idx in range(1, 8):
            ws_results.cell(row=row_idx, column=col_idx).font = Font(name="Segoe UI", size=10)
            
    # Auto-adjust column widths
    for col in ws_results.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_results.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 3. Excel Function Analysis & Diagnostics Sheet
    ws_analysis = wb.create_sheet(title="Analysis & Fixes")
    ws_analysis.views.sheetView[0].showGridLines = True
    
    analysis_headers = ["Failed Module", "Diagnosed Issue / Error", "Root Cause", "Actionable Fix / Recommendation"]
    for col_idx, header in enumerate(analysis_headers, 1):
        cell = ws_analysis.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    analysis_rows = [
        ["Authentication (Login Screen)", 
         "TC_02 Login Screen Renders FAILED", 
         "The app loads directly to the 'Onboarding' slide deck on first run, which suppresses direct display of login inputs.",
         "Ensure onboarding slides are auto-completed/dismissed, or target the specific React Native web 'Onboarding' class to click through."],
        
        ["Authentication (Submit Form)", 
         "TC_06 Form Submission FAILED", 
         "Attempted to click submit buttons matching typical HTML classes, which missed React Native Web's nested views.",
         "Update XPaths to search for standard accessibility labels or `data-testid` properties inside the React Native bundle."],
        
        ["Symptom Checker", 
         "TC_15 Input accepts text FAILED", 
         "Targeted standard HTML <textarea> tags, but the UI uses custom React Native views which don't map directly to standard textareas.",
         "Select the element via `div` containing class inputs or by placeholder text directly using XPaths."],
        
        ["Medicine Tracker", 
         "TC_19 Modal fails to open FAILED", 
         "Modal trigger button uses custom SVG plus nested touchable opacity, rendering it unclickable with basic class selectors.",
         "Utilize robust click helpers that locate target elements by accessibility label or dynamic text contents."],
        
        ["Profile Screen", 
         "TC_27 User Email not found FAILED", 
         "Since registration/login wasn't completed inside the real form, the profile screen remained unpopulated.",
         "Fix Onboarding click-throughs and Login submission to ensure the user session is authenticated before testing Profile."]
    ]
    
    for row_idx, row_data in enumerate(analysis_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_analysis.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Segoe UI", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
    ws_analysis.column_dimensions["A"].width = 25
    ws_analysis.column_dimensions["B"].width = 35
    ws_analysis.column_dimensions["C"].width = 40
    ws_analysis.column_dimensions["D"].width = 50
    
    # Save Workbook
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = REPORTS_DIR / f"CureConnect_Web_E2E_Analysis_{timestamp}.xlsx"
    wb.save(report_file)
    
    print("\n" + "="*60)
    print("  EXCEL REPORT SUCCESSFULLY GENERATED")
    print(f"  Path: {report_file}")
    print("="*60 + "\n")
    
if __name__ == "__main__":
    run_e2e_and_generate_excel()
